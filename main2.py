from datetime import datetime, time, timezone, timedelta
import math
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def to_utc(t, tz_offset):
    """Convert local time to UTC."""
    return t - timedelta(hours=tz_offset)

def main():

    tz1 = 1
    tz2 = 8

    tz1_sleep_start_ltz = 22.
    tz1_sleep_stop_ltz = 7.

    tz2_sleep_start_ltz = 22.
    tz2_sleep_stop_ltz = 7.

    travel_start_ltz = datetime(2025, 6, 1, 14, 30)
    travel_stop_ltz = datetime(2025, 6, 2, 14, 30)
    # translate all to UTC
    
    #END OF INPUT
    
    tz1_sleep_start = (tz1_sleep_start_ltz - tz1) % 24
    tz1_sleep_stop = (tz1_sleep_stop_ltz - tz1) % 24
    tz2_sleep_start = (tz2_sleep_start_ltz - tz2) % 24
    tz2_sleep_stop = (tz2_sleep_stop_ltz - tz2) % 24
    

    CBTmin_start = (tz1_sleep_start + 6) % 24
    CBTmin_target = (tz2_sleep_start + 6) % 24
    
    
    travel_start = travel_start_ltz + timedelta(hours=-tz1)
    travel_stop = travel_stop_ltz + timedelta(hours=-tz2)
    print("Travel start UTC:", travel_start)
    print("Travel stop UTC:", travel_stop)
    print("CBTmin start:", CBTmin_start)
    print("CBTmin target:", CBTmin_target)

    # CBTmin is the time of day when core body temperature is at its lowest.
    # CBTmin is usually around 6am, but can vary by a few hours depending on the individual and their sleep schedule.
    
    # lets start at travel start
    timetable = [] # list of dict

    current_CBTmin = CBTmin_start
    next_CBTmin = current_CBTmin
    print(current_CBTmin-CBTmin_target)
    current_iter_time = travel_start.replace(hour=0, minute=0, second=0, microsecond=0)

    while abs((current_CBTmin-CBTmin_target) % 24) > 0.5:
        print(current_CBTmin-CBTmin_target)
        tt_dict = {
            "t": current_iter_time,
            "sleep": False,
            "travel": False,
            "melatonin": False,
            "light_exposure": False,
            "light_avoidance": False,
            "CBTmin": False,
        }
        
        if current_iter_time >= travel_stop:
            current_sleep_start = tz2_sleep_start
            current_sleep_stop = tz2_sleep_stop
        else:
            current_sleep_start = tz1_sleep_start
            current_sleep_stop = tz1_sleep_stop

        print(math.floor(current_CBTmin), current_iter_time.hour)

        if math.floor(current_CBTmin) == current_iter_time.hour:
            tt_dict["CBTmin"] = True

        if travel_start <= current_iter_time <= travel_stop:
            tt_dict["sleep"] = False
            tt_dict["travel"] = True
        
        if current_sleep_start < current_sleep_stop:
            if current_sleep_start <= current_iter_time.hour < current_sleep_stop:
                tt_dict["sleep"] = True
        else:
            if current_iter_time.hour >= current_sleep_start or current_iter_time.hour < current_sleep_stop:
                tt_dict["sleep"] = True

        if 0 < ((current_CBTmin-CBTmin_target) % 24) < 12:
            # current is later than target. ie East
            best_melatonin = -11.5
            if math.floor((current_CBTmin + best_melatonin) % 24) == current_iter_time.hour:
                tt_dict["melatonin"] = True
                next_CBTmin = (next_CBTmin - 1) % 24
            
        elif 0 < ((CBTmin_target-current_CBTmin) % 24) < 12:
            # current is earlier than target. ie West
            best_melatonin = 4
            if math.floor((current_CBTmin + best_melatonin) % 24) == current_iter_time.hour:
                tt_dict["melatonin"] = True
                next_CBTmin = (next_CBTmin + 1) % 24
        
        if math.floor((current_CBTmin + 12) % 24) == current_iter_time.hour:
            current_CBTmin = next_CBTmin

        current_iter_time = current_iter_time + timedelta(hours=1)
        timetable.append(tt_dict)
    
    print(timetable)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Create headers (hours)
    for hour in range(24):
        ws.cell(row=1, column=hour + 2, value=f"{hour:02d}:00")
    
    # Group entries by date
    current_date = None
    row = 1
    
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    for entry in timetable:
        date = entry['t'].date()
        if date != current_date:
            row += 1
            current_date = date
            ws.cell(row=row, column=1, value=date.strftime('%Y-%m-%d'))
        
        hour = entry['t'].hour
        cell = ws.cell(row=row, column=hour + 2)
        
        # Add cell content and formatting
        content = ''
        if entry['sleep']:
            cell.fill = gray_fill
        if entry['travel']:
            content += 'T'
        if entry['melatonin']:
            content += 'M'
        if entry['CBTmin']:
            content += 'C'
        
        cell.value = content if content else None
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    for i in range(24):
        col_letter = ws.cell(row=1, column=i + 2).column_letter
        ws.column_dimensions[col_letter].width = 6
    
    wb.save("schedule.xlsx")


if __name__ == "__main__":
    main()